
import sys
from pathlib import Path
from openpyxl import load_workbook

def inspect_excel():
    file_path = Path("MonRep251201_00000_00200_11412.xlsx")
    print(f"Inspecting file: {file_path.absolute()}")
    
    if not file_path.exists():
        print("File does not exist!")
        return

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        print(f"Sheet Title: {ws.title}")
        print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
        
        print("\n--- Rows 1-100 ---")
        # Read first 100 rows
        for r in range(1, min(101, ws.max_row + 1)):
            row_vals = []
            # Read first 12 columns to be sure
            for c in range(1, 13):
                val = ws.cell(row=r, column=c).value
                # Simple formatting
                if val is None:
                    val_str = ""
                else:
                    val_str = str(val)
                row_vals.append(val_str)
            
            # Print only non-empty rows or interesting rows
            # if any(row_vals):
            print(f"Row {r:3d}: {row_vals}")
            
        wb.close()
        
    except Exception as e:
        print(f"Exception: {e}")

if __name__ == "__main__":
    inspect_excel()
