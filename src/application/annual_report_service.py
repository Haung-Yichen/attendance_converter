"""
Annual Report Service Module

Application layer service that orchestrates the generation of annual
attendance reports. Coordinates file discovery, per-month parsing
(reusing existing single-month logic), domain-level aggregation, and
infrastructure-level export.

Design:
- Depends on domain entities and infrastructure – never on PyQt.
- Receives all collaborators via constructor injection so that every
  dependency is explicit and replaceable in tests.
- Failures on individual months are collected as warnings, not raised.
"""

from __future__ import annotations

from calendar import monthrange
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import (
    Callable,
    Dict,
    List,
    Optional,
    Protocol,
    Set,
    Tuple,
)

from config.config_manager import AppConfig, ColorLogic, TimeRule
from domain.annual_aggregator import (
    AnnualAggregationResult,
    AnnualAggregator,
)
from domain.attendance_logic import AttendanceLogicFactory
from domain.entities import (
    MonthlyAttendance,
    Staff,
    StaffType,
)
from domain.rate_calculator import RateCalculator
from domain.sorting import sort_attendance_list
from domain.staff_classifier import StaffClassifier
from infrastructure.excel_parser import ExcelParser
from infrastructure.filename_parser import FilenameParser
from infrastructure.logger import get_logger

logger = get_logger("AnnualReportService")


# =============================================================================
# Parameter / Result DTOs
# =============================================================================

@dataclass
class AnnualReportParams:
    """
    All parameters needed to generate an annual report.

    The UI layer constructs this object and hands it to the service;
    the service never reaches back into UI or config globals.
    """
    year: int
    search_root: Path
    output_path: Path
    staff_csv_path: Path

    # Time rules
    internal_time_rule: TimeRule = field(default_factory=TimeRule)
    external_time_rule: TimeRule = field(default_factory=TimeRule)

    # Holidays
    holidays: Set[date] = field(default_factory=set)

    # Threshold
    rate_threshold: int = 80
    sort_by: str = "attendance_rate"

    # Color logic
    color_logic: Optional[ColorLogic] = None

    # Output options
    generate_pdf: bool = False


@dataclass
class AnnualReportResult:
    """Outcome of an annual report generation attempt."""
    success: bool
    year: int
    output_path: Path
    months_processed: List[int] = field(default_factory=list)
    months_missing: List[int] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    internal_count: int = 0
    external_count: int = 0
    error_message: str = ""


# =============================================================================
# Progress callback protocol
# =============================================================================

class ProgressCallback(Protocol):
    """Callable signature for progress updates sent back to the caller."""

    def __call__(self, current: int, total: int, message: str) -> None: ...


# =============================================================================
# Service
# =============================================================================

class AnnualReportService:
    """
    Application service for generating **annual** attendance reports.

    Workflow:
    1. Discover all 701Client files under ``search_root`` whose filename
       encodes a date belonging to the target year.
    2. For each discovered file, reuse the existing single-month
       ``ExcelParser`` → ``AttendanceLogicFactory`` → ``RateCalculator``
       pipeline to produce ``List[MonthlyAttendance]``.
    3. Pass the collected monthly data to ``AnnualAggregator`` (domain).
    4. Export the aggregated result via infrastructure writers.

    Dependencies are injected through the constructor so the service
    stays testable without touching the file system.
    """

    def __init__(
        self,
        *,
        parser_factory: Callable[[], ExcelParser] | None = None,
        classifier_factory: Callable[[], StaffClassifier] | None = None,
        aggregator_factory: Callable[[int], AnnualAggregator] | None = None,
    ) -> None:
        """
        Args:
            parser_factory: Factory that returns a fresh ``ExcelParser``.
                Defaults to ``ExcelParser``.
            classifier_factory: Factory that returns a fresh
                ``StaffClassifier``.  Defaults to ``StaffClassifier``.
            aggregator_factory: Factory ``(rate_threshold) -> AnnualAggregator``.
                Defaults to ``AnnualAggregator``.
        """
        self._parser_factory = parser_factory or ExcelParser
        self._classifier_factory = classifier_factory or StaffClassifier
        self._aggregator_factory = aggregator_factory or AnnualAggregator

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #

    def generate(
        self,
        params: AnnualReportParams,
        on_progress: ProgressCallback | None = None,
    ) -> AnnualReportResult:
        """
        Generate the annual attendance report.

        Args:
            params: All configuration packed into a DTO.
            on_progress: Optional callback ``(current, total, message)``
                invoked after each month is processed so that the UI can
                update a progress bar.

        Returns:
            ``AnnualReportResult`` containing outcome, warnings, and metadata.
        """
        warnings: List[str] = []

        # --- 1. Discover files ---------------------------------------- #
        files_by_month = self._discover_files(params.search_root, params.year)

        if not files_by_month:
            return AnnualReportResult(
                success=False,
                year=params.year,
                output_path=params.output_path,
                error_message=(
                    f"在 {params.search_root} 下找不到 {params.year} 年的 "
                    f"701Client 報表 (MonRepYYMMDD.xlsx)。"
                ),
            )

        total_months = len(files_by_month)
        logger.info(
            f"找到 {total_months} 個月份的來源檔案 "
            f"(月份: {sorted(files_by_month.keys())})"
        )

        # --- 2. Load staff classification ----------------------------- #
        classifier = self._classifier_factory()
        classifier.load_from_csv(params.staff_csv_path)

        # --- 3. Parse each month -------------------------------------- #
        monthly_data: Dict[int, List[MonthlyAttendance]] = {}

        for idx, (month, file_path) in enumerate(
            sorted(files_by_month.items()), start=1
        ):
            if on_progress:
                on_progress(
                    idx, total_months,
                    f"正在處理 {params.year}/{month:02d} … ({file_path.name})",
                )

            try:
                month_attendances = self._parse_single_month(
                    file_path=file_path,
                    year=params.year,
                    month=month,
                    classifier=classifier,
                    params=params,
                )
                if month_attendances:
                    monthly_data[month] = month_attendances
                else:
                    msg = f"{params.year}/{month:02d}: 檔案解析後無有效資料 ({file_path.name})"
                    logger.warning(msg)
                    warnings.append(msg)
            except Exception as exc:  # noqa: BLE001
                msg = f"{params.year}/{month:02d}: 解析失敗 – {exc}"
                logger.warning(msg)
                warnings.append(msg)

        if not monthly_data:
            return AnnualReportResult(
                success=False,
                year=params.year,
                output_path=params.output_path,
                warnings=warnings,
                error_message="所有月份的來源檔案都無法成功解析，無法產生年度報表。",
            )

        # --- 4. Aggregate --------------------------------------------- #
        aggregator = self._aggregator_factory(params.rate_threshold)
        result = aggregator.aggregate(params.year, monthly_data)

        # --- 5. Export ------------------------------------------------ #
        try:
            self._export(result, params)
        except Exception as exc:  # noqa: BLE001
            return AnnualReportResult(
                success=False,
                year=params.year,
                output_path=params.output_path,
                months_processed=result.available_months,
                months_missing=result.missing_months,
                warnings=warnings,
                error_message=f"匯出年度報表失敗: {exc}",
            )

        return AnnualReportResult(
            success=True,
            year=params.year,
            output_path=params.output_path,
            months_processed=result.available_months,
            months_missing=result.missing_months,
            warnings=warnings,
            internal_count=len(result.internal_summaries),
            external_count=len(result.external_summaries),
        )

    # ------------------------------------------------------------------ #
    # Static helper – construct params from AppConfig
    # ------------------------------------------------------------------ #

    @staticmethod
    def build_params(
        config: AppConfig,
        year: int,
        search_root: Path,
        output_path: Path,
    ) -> AnnualReportParams:
        """
        Convenience builder that maps ``AppConfig`` → ``AnnualReportParams``.

        Keeps the UI layer thin – it only needs to provide the year and
        root directory.
        """
        holidays_set: Set[date] = set()
        for date_str in config.holidays.custom_dates:
            try:
                parts = date_str.split("-")
                holidays_set.add(
                    date(int(parts[0]), int(parts[1]), int(parts[2]))
                )
            except (ValueError, IndexError):
                pass

        return AnnualReportParams(
            year=year,
            search_root=search_root,
            output_path=output_path,
            staff_csv_path=Path(config.paths.staff_csv),
            internal_time_rule=config.time_rules.internal,
            external_time_rule=config.time_rules.external,
            holidays=holidays_set,
            rate_threshold=config.ui_prefs.rate_threshold,
            sort_by=config.output_settings.sort_by,
            color_logic=config.ui_prefs.color_logic,
        )

    # ------------------------------------------------------------------ #
    # Private: file discovery
    # ------------------------------------------------------------------ #

    @staticmethod
    def _discover_files(
        search_root: Path,
        year: int,
    ) -> Dict[int, Path]:
        """
        Scan *search_root* (non-recursively) for 701Client export files
        whose filename encodes a date in *year*.

        Resolution order for each file:
        1. Try the ``_YYYYMM`` **suffix** (data month) first.
        2. Fall back to the ``MonRepYYMMDD`` **prefix** (export date).

        Returns a dict mapping ``month`` → ``Path``.  If multiple files
        share the same month, the **last** one found (alphabetically
        latest, i.e. newest export date) wins.
        """
        files_by_month: Dict[int, Path] = {}

        if not search_root.is_dir():
            logger.warning(f"搜尋路徑不存在或非目錄: {search_root}")
            return files_by_month

        for entry in sorted(search_root.iterdir()):
            if not entry.is_file():
                continue
            if entry.suffix.lower() not in (".xlsx", ".xls"):
                continue

            # Prefer the trailing _YYYYMM suffix (= actual data month)
            parsed = FilenameParser.try_parse_data_month(entry.name)
            if parsed is None:
                # Fall back to the MonRepYYMMDD prefix (= export date)
                parsed = FilenameParser.try_parse_report_date(entry.name)
            if parsed is None:
                continue

            file_year, file_month = parsed
            if file_year == year:
                files_by_month[file_month] = entry
                logger.debug(f"發現 {year}/{file_month:02d} 來源: {entry.name}")

        return files_by_month

    # ------------------------------------------------------------------ #
    # Private: single-month parsing (reuses existing pipeline)
    # ------------------------------------------------------------------ #

    def _parse_single_month(
        self,
        file_path: Path,
        year: int,
        month: int,
        classifier: StaffClassifier,
        params: AnnualReportParams,
    ) -> List[MonthlyAttendance]:
        """
        Parse one monthly file and return per-employee
        ``MonthlyAttendance`` objects.

        The *month* parameter is the **expected** month derived from the
        filename.  After parsing we verify against the actual dates in
        the data and use those instead (the filename may encode the
        export date, not the data month).

        Employees not found in the classifier are **skipped** with a
        logger warning (not raised), to keep the annual pipeline robust.
        """
        parser = self._parser_factory()

        # Pass the year hint so that MM/DD-only dates are resolved
        parsed_prefix = FilenameParser.try_parse_report_date(file_path.name)
        year_hint = parsed_prefix[0] if parsed_prefix else year
        raw_data = parser.parse_file(file_path, year=year_hint)

        if not raw_data:
            return []

        # Determine the ACTUAL data year/month from the records
        # (mirrors the approach in report_service.py)
        first_date = raw_data[0].date
        actual_year = first_date.year
        actual_month = first_date.month

        if actual_month != month:
            logger.debug(
                f"檔案 {file_path.name}: 預期月份={month}, "
                f"實際資料月份={actual_month} — 以實際資料為準"
            )

        records_by_name = parser.get_records_by_month(actual_year, actual_month)
        if not records_by_name:
            return []

        # Override month/year for downstream calculations
        month = actual_month
        year = actual_year

        _, num_days = monthrange(year, month)
        rate_calc = RateCalculator(holidays=params.holidays)
        result: List[MonthlyAttendance] = []

        for name, raw_rows in records_by_name.items():
            staff = classifier.get_staff_by_name(name)
            if staff is None:
                logger.debug(
                    f"[{year}/{month:02d}] 人員 '{name}' 未在名單中，已略過"
                )
                continue

            records = parser.convert_to_attendance_records(raw_rows)

            # Apply attendance logic
            strategy = AttendanceLogicFactory.get_strategy(staff.staff_type)
            time_rule = (
                params.internal_time_rule
                if staff.staff_type == StaffType.INTERNAL
                else params.external_time_rule
            )
            for record in records:
                record.status = strategy.determine_status(record, time_rule)
                record.remark = strategy.get_remark(record, time_rule)

            # Build work-day set for this month
            work_days_set: Set[int] = set()
            for day in range(1, num_days + 1):
                d = date(year, month, day)
                if staff.should_work_on(d) and d not in params.holidays:
                    work_days_set.add(day)

            monthly = rate_calc.calculate_monthly_attendance(
                staff,
                records,
                year,
                month,
                params.rate_threshold,
                work_days=work_days_set,
            )
            result.append(monthly)

        return result

    # ------------------------------------------------------------------ #
    # Private: export
    # ------------------------------------------------------------------ #

    def _export(
        self,
        aggregation: AnnualAggregationResult,
        params: AnnualReportParams,
    ) -> None:
        """
        Export the aggregated annual data to an Excel file.

        This is a **skeleton** – the full formatting can be expanded
        later using a dedicated ``AnnualExcelWriter``.  For now we
        create a simple openpyxl workbook with the summary data.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = f"{params.year}年度出席率報表"

        # ----- Header row ----- #
        headers = ["姓名", "類別", "狀態"]
        for m in range(1, 13):
            headers.append(f"{m}月")
        headers += ["應出席(加總)", "實際出席(加總)", "年度出席率"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_align = Alignment(horizontal="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # ----- Data rows ----- #
        row = 2
        all_summaries = (
            aggregation.internal_summaries + aggregation.external_summaries
        )

        status_label = {
            "ACTIVE": "在職",
            "RESIGNED": "離職",
            "NEW_HIRE": "新進",
        }

        rate_fills = {
            "RED": PatternFill("solid", fgColor="FF6B6B"),
            "YELLOW": PatternFill("solid", fgColor="FFD700"),
            "GREEN": PatternFill("solid", fgColor="90EE90"),
        }

        for summary in all_summaries:
            ws.cell(row=row, column=1, value=summary.staff.name)
            type_label = "內勤" if summary.staff.staff_type == StaffType.INTERNAL else "外勤"
            ws.cell(row=row, column=2, value=type_label)
            ws.cell(
                row=row, column=3,
                value=status_label.get(summary.status.name, summary.status.name),
            )

            for m in range(1, 13):
                col = 3 + m  # columns 4-15
                snap = summary.monthly_snapshots.get(m)
                if snap is not None:
                    cell = ws.cell(
                        row=row, column=col,
                        value=round(snap.attendance_rate, 1),
                    )
                    cell.alignment = Alignment(horizontal="center")
                    fill = rate_fills.get(snap.rate_color.name)
                    if fill:
                        cell.fill = fill
                else:
                    ws.cell(row=row, column=col, value="—").alignment = Alignment(
                        horizontal="center"
                    )

            ws.cell(row=row, column=16, value=summary.total_required_days)
            ws.cell(row=row, column=17, value=summary.total_actual_days)

            rate_cell = ws.cell(
                row=row, column=18,
                value=round(summary.annual_attendance_rate, 1),
            )
            rate_cell.alignment = Alignment(horizontal="center")
            fill = rate_fills.get(summary.rate_color.name)
            if fill:
                rate_cell.fill = fill

            row += 1

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = 12

        # Name column wider
        ws.column_dimensions["A"].width = 16

        # Save
        params.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(params.output_path)
        logger.info(f"年度報表已儲存: {params.output_path}")
