"""
Excel Parser Module

Handles parsing and cleaning raw Excel attendance exports.
Cleans messy data like "Name[]" patterns and "*" symbols.
"""

import re
from datetime import datetime, time, date
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain.entities import AttendanceRecord, AttendanceStatus
from infrastructure.logger import get_logger

logger = get_logger("ExcelParser")


# ==============================================================================
# Custom Exceptions
# ==============================================================================
class AttendanceError(Exception):
    """Base exception for attendance-related errors."""
    pass


class ExcelFormatError(AttendanceError):
    """Raised when the Excel file format is unrecognized or invalid."""
    pass


class UnclassifiedStaffError(AttendanceError):
    """
    Raised when a staff member is not found in the classification list.
    
    The UI layer should catch this and prompt the user for classification.
    """
    def __init__(self, name: str, message: str = None):
        self.staff_name = name
        self.message = message or f"人員 '{name}' 未在人員名單中分類，請指定類別（內勤/外勤）。"
        super().__init__(self.message)


# ==============================================================================
# Data Classes
# ==============================================================================
@dataclass
class RawAttendanceRow:
    """Raw attendance data from Excel."""
    name: str
    date: date
    check_in: Optional[time]
    check_out: Optional[time]


# ==============================================================================
# ExcelParser Class
# ==============================================================================
class ExcelParser:
    """
    Parses raw Excel attendance exports.
    
    Handles:
    - Cleaning "Name[]" garbage patterns
    - Removing "*" symbols from time values
    - Extracting attendance data into domain entities
    """
    
    # Pattern to clean name fields like "John[]" or "Mary[*]"
    NAME_CLEAN_PATTERN = re.compile(r'\[.*?\]')
    
    # Pattern to clean time values with asterisks
    TIME_CLEAN_PATTERN = re.compile(r'\*')
    
    # Maximum rows to search for header
    MAX_HEADER_SEARCH_ROWS = 15
    
    def __init__(self):
        self._raw_data: List[RawAttendanceRow] = []
        self._unique_names: List[str] = []
        self._year: Optional[int] = None
    
    def parse_file(self, file_path: Path, year: Optional[int] = None) -> List[RawAttendanceRow]:
        """
        Parse an Excel file and extract attendance data.
        
        Args:
            file_path: Path to the Excel file
            year: Optional year to use for date parsing (for MM/DD formats)
            
        Returns:
            List of RawAttendanceRow objects
            
        Raises:
            ExcelFormatError: If the file format cannot be recognized
        """
        self._raw_data = []
        self._unique_names = []
        self._year = year
        
        if not file_path.exists():
            logger.warning(f"來源檔案不存在: {file_path}")
            return []
        
        logger.info(f"開始解析 Excel 檔案: {file_path.name}")
        
        wb = load_workbook(file_path, data_only=True)
        
        # Parse ALL worksheets, not just the active one
        for ws in wb.worksheets:
            try:
                sheet_data = self._parse_worksheet(ws)
                self._raw_data.extend(sheet_data)
            except ExcelFormatError:
                # Re-raise format errors
                raise
            except Exception as e:
                # Log and skip sheets that fail to parse
                logger.warning(f"解析工作表 '{ws.title}' 時發生錯誤，已跳過: {e}")
                continue
        
        # Extract unique names
        seen_names = set()
        for row in self._raw_data:
            if row.name not in seen_names:
                self._unique_names.append(row.name)
                seen_names.add(row.name)
        
        wb.close()
        logger.info(f"解析完成: 共 {len(self._raw_data)} 筆記錄, {len(self._unique_names)} 位人員")
        return self._raw_data
    
    def _parse_worksheet(self, ws: Worksheet) -> List[RawAttendanceRow]:
        """Parse a worksheet and extract attendance rows.
        
        Expected format (MonRep export):
        - Column A (1): 部門 (Department)
        - Column B (2): 姓名 (Name)
        - Column C (3): 日期 (Date)
        - Column D (4): 遲到 (Late)
        - Column E (5): 早退 (Early leave)
        - Column F (6): 加班 (Overtime)
        - Column G (7): 工時 (Work hours)
        - Column H (8): 假別(1) / 上班 (Check-in)
        - Column I (9): 假別(2) / 下班 (Check-out)
        
        Raises:
            ExcelFormatError: If header row cannot be found within MAX_HEADER_SEARCH_ROWS
        """
        rows = []
        
        # Detect header row by looking for column headers
        header_row = None
        name_col = None
        date_col = None
        check_in_col = None  # Must be detected
        check_out_col = None  # Must be detected
        
        # Search for header row in first MAX_HEADER_SEARCH_ROWS rows
        max_search_rows = min(self.MAX_HEADER_SEARCH_ROWS, ws.max_row + 1)
        max_search_cols = min(15, ws.max_column + 1)
        
        for row_idx in range(1, max_search_rows + 1):
            for col_idx in range(1, max_search_cols + 1):
                cell_value = str(ws.cell(row_idx, col_idx).value or '').strip()
                cell_lower = cell_value.lower()
                
                # Detect name column
                if any(keyword in cell_lower for keyword in ['name', '姓名', '員工']):
                    header_row = row_idx
                    name_col = col_idx
                
                # Detect date column
                if any(keyword in cell_lower for keyword in ['date', '日期']):
                    date_col = col_idx
                
                # Detect check-in column (上班)
                if '上班' in cell_value:
                    check_in_col = col_idx
                
                # Detect check-out column (下班)
                if '下班' in cell_value:
                    check_out_col = col_idx
        
        # CRITICAL: Check-in and Check-out columns MUST be found
        if check_in_col is None or check_out_col is None:
            missing_cols = []
            if check_in_col is None:
                missing_cols.append("'上班'")
            if check_out_col is None:
                missing_cols.append("'下班'")
            raise ExcelFormatError(
                f"無法識別工作表 '{ws.title}' 的打卡欄位。\n"
                f"在前 {self.MAX_HEADER_SEARCH_ROWS} 列中找不到 {' 和 '.join(missing_cols)} 欄位。\n"
                f"無法計算出席率，請確認 Excel 檔案格式是否正確。"
            )
        
        # If no standard header found for name/date, use defaults (MonRep format)
        # In MonRep exports, each worksheet represents one person (sheet name = person name)
        use_sheet_name_as_person = False
        
        if header_row is None or name_col is None or date_col is None:
            logger.debug(
                f"工作表 '{ws.title}': 未找到標準表頭，使用預設欄位配置 "
                f"(姓名=B, 日期=C)"
            )
            # Use defaults - this is the MonRep format where each sheet is a person
            header_row = 1
            name_col = 2  # B column
            date_col = 3  # C column
            
            # Use sheet title as the person's name (common in MonRep exports)
            use_sheet_name_as_person = True
        
        logger.debug(
            f"工作表 '{ws.title}': 表頭列={header_row}, "
            f"姓名欄={name_col}, 日期欄={date_col}, "
            f"上班欄={check_in_col}, 下班欄={check_out_col}"
        )
        
        # Parse data rows - track name across rows since name only appears once per person
        # If using sheet name as person, extract clean name from sheet title
        current_name = None
        if use_sheet_name_as_person:
            # Clean sheet title: remove trailing '-' and other garbage
            sheet_name = ws.title.rstrip('-').strip()
            if sheet_name:
                current_name = self._clean_name(sheet_name)
        skipped_rows = 0
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                name_cell = ws.cell(row_idx, name_col).value
                
                # Update current name if we find a new one
                if name_cell:
                    current_name = self._clean_name(str(name_cell))
                
                # Skip if no name established yet
                if not current_name:
                    continue
                
                # Extract date - skip non-date rows (summary rows at bottom)
                date_val = self._extract_date(ws.cell(row_idx, date_col).value, self._year)
                if not date_val:
                    continue
                
                # Extract times from detected columns
                check_in = self._extract_time(ws.cell(row_idx, check_in_col).value)
                check_out = self._extract_time(ws.cell(row_idx, check_out_col).value)
                
                rows.append(RawAttendanceRow(
                    name=current_name,
                    date=date_val,
                    check_in=check_in,
                    check_out=check_out
                ))
                
            except ValueError as e:
                # Log the error and skip this row
                skipped_rows += 1
                logger.warning(
                    f"工作表 '{ws.title}' 第 {row_idx} 列解析失敗，已跳過: {e}"
                )
                continue
            except Exception as e:
                # Catch any other unexpected errors
                skipped_rows += 1
                logger.warning(
                    f"工作表 '{ws.title}' 第 {row_idx} 列發生非預期錯誤，已跳過: {e}"
                )
                continue
        
        if skipped_rows > 0:
            logger.info(f"工作表 '{ws.title}': 共跳過 {skipped_rows} 列有問題的資料")
        
        return rows
    
    def _clean_name(self, name: str) -> str:
        """Clean a name field by removing garbage patterns."""
        if not name:
            return ""
        # Remove [*] or [] patterns
        cleaned = self.NAME_CLEAN_PATTERN.sub('', name)
        return cleaned.strip()
    
    def _extract_date(self, value, year: Optional[int] = None) -> Optional[date]:
        """Extract date from a cell value.
        
        Handles formats like:
        - 12/01(一) - MM/DD with weekday in parentheses
        - 2024-12-01
        - 2024/12/01
        """
        if value is None:
            return None
        
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        
        # Try parsing string formats
        str_val = str(value).strip()
        
        # Handle MM/DD(weekday) format like "12/01(一)" or "12/05 (五 *)"
        # Remove the weekday part in parentheses including any spaces/asterisks
        weekday_pattern = re.compile(r'\s*\([一二三四五六日月火水木金土]\s*\*?\)')
        str_val_clean = weekday_pattern.sub('', str_val).strip()
        
        # Determine year - warn if we have to fallback to current year
        if year is None:
            year = datetime.now().year
            # Only warn for MM/DD format (short date strings)
            if '/' in str_val_clean and len(str_val_clean) <= 5:
                logger.warning(
                    f"日期 '{str_val}' 沒有年份資訊，使用當前年份 {year}。"
                    f"如果資料跨年份，可能會產生錯誤結果。"
                )
        
        # Try MM/DD format (common in Taiwan/Japan)
        if '/' in str_val_clean and len(str_val_clean) <= 5:
            try:
                parts = str_val_clean.split('/')
                if len(parts) == 2:
                    month = int(parts[0])
                    day = int(parts[1])
                    return date(year, month, day)
            except (ValueError, IndexError):
                pass
        
        # Try standard formats
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(str_val_clean, fmt).date()
            except ValueError:
                continue
        
        return None
    
    def _extract_time(self, value) -> Optional[time]:
        """Extract time from a cell value, cleaning asterisks."""
        if value is None:
            return None
        
        if isinstance(value, datetime):
            return value.time()
        if isinstance(value, time):
            return value
        
        # Clean and parse string
        str_val = self.TIME_CLEAN_PATTERN.sub('', str(value)).strip()
        
        if not str_val:
            return None
        
        for fmt in ['%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M:%S %p']:
            try:
                return datetime.strptime(str_val, fmt).time()
            except ValueError:
                continue
        
        # If all formats fail, don't raise - just return None and log
        logger.debug(f"無法解析時間格式: '{value}'")
        return None
    
    def get_unique_names(self) -> List[str]:
        """Get list of unique staff names found in the file."""
        return self._unique_names
    
    def get_records_for_name(self, name: str) -> List[RawAttendanceRow]:
        """Get all records for a specific staff member."""
        return [row for row in self._raw_data if row.name == name]
    
    def get_records_by_month(
        self, 
        year: int, 
        month: int
    ) -> Dict[str, List[RawAttendanceRow]]:
        """
        Get records grouped by name for a specific month.
        
        Args:
            year: Year to filter
            month: Month to filter (1-12)
            
        Returns:
            Dictionary mapping names to their records
        """
        result: Dict[str, List[RawAttendanceRow]] = {}
        
        for row in self._raw_data:
            if row.date.year == year and row.date.month == month:
                if row.name not in result:
                    result[row.name] = []
                result[row.name].append(row)
        
        return result
    
    def convert_to_attendance_records(
        self,
        raw_rows: List[RawAttendanceRow]
    ) -> List[AttendanceRecord]:
        """
        Convert raw rows to domain AttendanceRecord objects.
        
        Args:
            raw_rows: List of raw attendance rows
            
        Returns:
            List of AttendanceRecord objects
        """
        records = []
        for raw in raw_rows:
            status = AttendanceStatus.ABSENT
            if raw.check_in or raw.check_out:
                status = AttendanceStatus.NORMAL  # Will be refined by logic layer
            
            records.append(AttendanceRecord(
                date=raw.date,
                check_in=raw.check_in,
                check_out=raw.check_out,
                status=status
            ))
        
        return records
