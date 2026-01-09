"""
Excel Writer Module

Generates formatted Excel attendance reports with styling.
Applies color formatting based on business rules.
"""

from datetime import date, time
from calendar import monthrange
from pathlib import Path
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from domain.entities import (
    Staff, StaffType, AttendanceRecord, AttendanceStatus,
    MonthlyAttendance, RateColorTier
)
from config.config_manager import ColorLogic, TimeRule


class ExcelWriter:
    """
    Generates formatted Excel attendance reports.
    
    Output format:
    - Row 1: Day numbers (1-31) + Remarks column at end
    - Row 2 (per user): Merged Name | In-Time cells
    - Row 3 (per user): Merged Name | Out-Time cells
    
    Styling:
    - Green/Red backgrounds based on color_logic settings
    - Gray columns for non-work days (External: Tue/Thu/Sat/Sun)
    - 3-tier color for attendance rates
    """
    
    # Color definitions (支援所有設定中可用的顏色)
    COLORS = {
        'green': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
        'red': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
        'yellow': PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid'),
        'orange': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
        'blue': PatternFill(start_color='6B8CFF', end_color='6B8CFF', fill_type='solid'),
        'purple': PatternFill(start_color='DDA0DD', end_color='DDA0DD', fill_type='solid'),
        'pink': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),
        'black': PatternFill(start_color='333333', end_color='333333', fill_type='solid'),
        'gray': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'),
        'header': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
    }
    
    # Hex 色碼到顏色名稱的映射 (用於處理 config 中的 hex 格式)
    HEX_TO_NAME = {
        '#90EE90': 'green', '#90ee90': 'green',
        '#FF6B6B': 'red', '#ff6b6b': 'red',
        '#FFD700': 'yellow', '#ffd700': 'yellow',
        '#FFA500': 'orange', '#ffa500': 'orange',
        '#6B8CFF': 'blue', '#6b8cff': 'blue',
        '#DDA0DD': 'purple', '#dda0dd': 'purple',
        '#FFB6C1': 'pink', '#ffb6c1': 'pink',
        '#333333': 'black',
        '#D3D3D3': 'gray', '#d3d3d3': 'gray',
    }
    
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 粗邊框用於每個人的外框
    THICK_SIDE = Side(style='medium')
    THIN_SIDE = Side(style='thin')
    
    def __init__(self, color_logic: ColorLogic = None, time_rule: TimeRule = None):
        self.color_logic = color_logic or ColorLogic()
        self.time_rule = time_rule or TimeRule()
        self.wb: Optional[Workbook] = None
    
    def _get_fill(self, color_value: str) -> Optional[PatternFill]:
        """Get PatternFill from color value (name or hex code).
        
        Args:
            color_value: Color name ('green', 'red') or hex code ('#90EE90')
            
        Returns:
            PatternFill object or None if color is invalid/none/transparent
        """
        if not color_value or color_value in ('none', 'transparent'):
            return None
        
        # 直接是顏色名稱
        if color_value in self.COLORS:
            return self.COLORS[color_value]
        
        # 如果是 hex 碼，嘗試轉換為顏色名稱
        if color_value in self.HEX_TO_NAME:
            color_name = self.HEX_TO_NAME[color_value]
            return self.COLORS[color_name]
        
        return None
    
    def _is_dark_color(self, color_value: str) -> bool:
        """Check if color is dark (needs white text)."""
        if color_value in ('black', 'blue', 'purple'):
            return True
        # 處理 hex 格式
        if color_value in self.HEX_TO_NAME:
            return self.HEX_TO_NAME[color_value] in ('black', 'blue', 'purple')
        return False
    
    def create_report(
        self,
        internal_attendance: List[MonthlyAttendance],
        external_attendance: List[MonthlyAttendance],
        year: int,
        month: int,
        output_path: Path,
        holidays: set = None
    ) -> Path:
        """
        Create a complete attendance report with separate sheets.
        
        Args:
            internal_attendance: List of internal staff attendance
            external_attendance: List of external staff attendance
            year: Report year
            month: Report month
            output_path: Path to save the Excel file
            holidays: Set of holiday dates to exclude from columns
            
        Returns:
            Path to the created file
        """
        self.wb = Workbook()
        
        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)
        
        # Create Internal sheet
        internal_ws = self.wb.create_sheet("內勤出勤表")
        self._write_sheet(
            internal_ws, 
            internal_attendance, 
            year, 
            month,
            is_external=False,
            holidays=holidays
        )
        
        # Create External sheet
        external_ws = self.wb.create_sheet("外勤出勤表")
        self._write_sheet(
            external_ws, 
            external_attendance, 
            year, 
            month,
            is_external=True,
            holidays=holidays
        )
        
        # Save
        self.wb.save(output_path)
        return output_path
    
    def _write_sheet(
        self,
        ws,
        attendance_list: List[MonthlyAttendance],
        year: int,
        month: int,
        is_external: bool = False,
        holidays: set = None
    ):
        """Write attendance data to a worksheet.
        
        Only draws columns for work days:
        - Internal: Mon-Fri (weekday 0-4), excluding holidays
        - External: Mon/Wed/Fri (weekday 0, 2, 4), excluding holidays
        """
        _, num_days = monthrange(year, month)
        holidays = holidays or set()
        
        # Determine work days based on staff type
        if is_external:
            # External: Mon(0), Wed(2), Fri(4)
            work_weekdays = {0, 2, 4}
        else:
            # Internal: Mon-Fri (0-4)
            work_weekdays = {0, 1, 2, 3, 4}
        
        # Build list of work days for this month
        work_days = []
        for day in range(1, num_days + 1):
            d = date(year, month, day)
            if d.weekday() in work_weekdays and d not in holidays:
                work_days.append(day)
        
        # Chinese weekday names
        weekday_names = ['一', '二', '三', '四', '五', '六', '日']
        
        # Create day-to-column mapping (day -> col index starting from 2)
        day_to_col = {day: col + 2 for col, day in enumerate(work_days)}
        num_work_days = len(work_days)
        
        # Header row - Name column
        ws.cell(1, 1, "姓名").font = Font(bold=True, color='FFFFFF')
        ws.cell(1, 1).fill = self.COLORS['header']
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(1, 1).border = self.BORDER
        
        # Header row - Date columns (only work days)
        for day in work_days:
            col = day_to_col[day]
            d = date(year, month, day)
            weekday_str = weekday_names[d.weekday()]
            date_label = f"{month:02d}/{day:02d}({weekday_str})"
            
            cell = ws.cell(1, col, date_label)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = self.COLORS['header']
            cell.alignment = Alignment(horizontal='center', text_rotation=90)
            cell.border = self.BORDER
        
        # Summary columns after work days
        remarks_col = num_work_days + 2
        actual_col = num_work_days + 3
        rate_col = num_work_days + 4
        
        # Remarks header
        cell = ws.cell(1, remarks_col, "備註")
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = self.COLORS['header']
        cell.alignment = Alignment(horizontal='center')
        cell.border = self.BORDER
        
        # Actual attendance days header
        cell = ws.cell(1, actual_col, "實際出勤天數")
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = self.COLORS['header']
        cell.alignment = Alignment(horizontal='center')
        cell.border = self.BORDER
        
        # Attendance rate header
        cell = ws.cell(1, rate_col, "出席率")
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = self.COLORS['header']
        cell.alignment = Alignment(horizontal='center')
        cell.border = self.BORDER
        
        # Data rows (2 rows per person: check-in and check-out)
        current_row = 2
        
        for monthly in attendance_list:
            staff = monthly.staff
            records_by_day = {r.date.day: r for r in monthly.records}
            
            in_row = current_row
            out_row = current_row + 1
            
            # Name cell (merged)
            ws.merge_cells(start_row=in_row, start_column=1, end_row=out_row, end_column=1)
            name_cell = ws.cell(in_row, 1, staff.name)
            name_cell.font = Font(bold=True)
            name_cell.alignment = Alignment(horizontal='center', vertical='center')
            name_cell.border = self.BORDER
            ws.cell(out_row, 1).border = self.BORDER
            
            # Time cells for each work day only
            for day in work_days:
                col = day_to_col[day]
                record = records_by_day.get(day)
                
                in_cell = ws.cell(in_row, col)
                out_cell = ws.cell(out_row, col)
                
                in_cell.border = self.BORDER
                out_cell.border = self.BORDER
                in_cell.alignment = Alignment(horizontal='center')
                out_cell.alignment = Alignment(horizontal='center')
                
                if record:
                    has_in = record.check_in is not None
                    has_out = record.check_out is not None
                    
                    if has_in and has_out:
                        # 兩個都有打卡，正常顯示時間
                        in_cell.value = record.check_in.strftime('%H:%M')
                        out_cell.value = record.check_out.strftime('%H:%M')
                        self._apply_status_colors(in_cell, out_cell, record)
                    
                    elif has_in and not has_out:
                        # 有打上班、沒打下班 → 缺少下班打卡紀錄
                        in_cell.value = record.check_in.strftime('%H:%M')
                        out_cell.value = self.color_logic.missing_punch_text
                        self._apply_status_colors(in_cell, out_cell, record)
                        self._apply_missing_punch_color(out_cell)
                    
                    elif not has_in and has_out:
                        # 沒打上班、有打下班 → 缺少上班打卡紀錄
                        in_cell.value = self.color_logic.missing_punch_text
                        out_cell.value = record.check_out.strftime('%H:%M')
                        self._apply_missing_punch_color(in_cell)
                        self._apply_status_colors(in_cell, out_cell, record)
                    
                    else:
                        # 兩個都沒打 (record 存在但無打卡) → 曠職
                        in_cell.value = self.color_logic.absent_text
                        out_cell.value = self.color_logic.absent_text
                        self._apply_absent_color(in_cell)
                        self._apply_absent_color(out_cell)
                else:
                    # 完全沒有 record → 曠職
                    in_cell.value = self.color_logic.absent_text
                    out_cell.value = self.color_logic.absent_text
                    self._apply_absent_color(in_cell)
                    self._apply_absent_color(out_cell)
            
            # Generate remarks based on status (only for work days)
            remark_items = []
            for day in work_days:
                record = records_by_day.get(day)
                if not record:
                    continue
                if record.status == AttendanceStatus.LATE:
                    remark_items.append(f"{record.date.day}日遲到")
                elif record.status == AttendanceStatus.EARLY_LEAVE:
                    remark_items.append(f"{record.date.day}日早退")
                if record.check_out and record.remark == "下班延遲打卡":
                    remark_items.append(f"{record.date.day}日下班延遲打卡")
            
            # Remarks cell
            remark_cell = ws.cell(in_row, remarks_col, ", ".join(remark_items) if remark_items else "")
            remark_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            remark_cell.border = self.BORDER
            ws.cell(out_row, remarks_col).border = self.BORDER
            ws.merge_cells(start_row=in_row, start_column=remarks_col, end_row=out_row, end_column=remarks_col)
            
            # Actual attendance days cell
            actual_cell = ws.cell(in_row, actual_col, str(monthly.actual_days))
            actual_cell.alignment = Alignment(horizontal='center', vertical='center')
            actual_cell.border = self.BORDER
            ws.cell(out_row, actual_col).border = self.BORDER
            ws.merge_cells(start_row=in_row, start_column=actual_col, end_row=out_row, end_column=actual_col)
            
            # Attendance rate cell
            rate_cell = ws.cell(in_row, rate_col, f"{monthly.attendance_rate:.1f}%")
            rate_cell.alignment = Alignment(horizontal='center', vertical='center')
            rate_cell.border = self.BORDER
            ws.cell(out_row, rate_col).border = self.BORDER
            ws.merge_cells(start_row=in_row, start_column=rate_col, end_row=out_row, end_column=rate_col)
            
            # Apply rate color
            if monthly.rate_color == RateColorTier.GREEN:
                rate_cell.fill = self.COLORS['green']
            elif monthly.rate_color == RateColorTier.YELLOW:
                rate_cell.fill = self.COLORS['yellow']
            else:
                rate_cell.fill = self.COLORS['red']
            
            # 為每個人的兩列套用粗外框
            self._apply_person_border(ws, in_row, out_row, rate_col)
            
            current_row += 2
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        for col in range(2, num_work_days + 2):
            ws.column_dimensions[get_column_letter(col)].width = 7
        ws.column_dimensions[get_column_letter(remarks_col)].width = 20
        ws.column_dimensions[get_column_letter(actual_col)].width = 14
        ws.column_dimensions[get_column_letter(rate_col)].width = 8
        
        # Set header row height for rotated date text
        ws.row_dimensions[1].height = 50
    
    def _apply_status_colors(self, in_cell, out_cell, record: AttendanceRecord):
        """Apply colors based on attendance status and color_logic settings."""
        if record.status == AttendanceStatus.NORMAL:
            # 使用新的顏色字串屬性
            if record.check_in:
                fill = self._get_fill(self.color_logic.normal_in_color)
                if fill:
                    in_cell.fill = fill
            if record.check_out:
                fill = self._get_fill(self.color_logic.normal_out_color)
                if fill:
                    out_cell.fill = fill
        
        elif record.status == AttendanceStatus.LATE:
            # 上班遲到用異常顏色，下班正常用正常顏色
            fill = self._get_fill(self.color_logic.abnormal_in_color)
            if fill:
                in_cell.fill = fill
            if record.check_out:
                fill = self._get_fill(self.color_logic.normal_out_color)
                if fill:
                    out_cell.fill = fill
        
        elif record.status == AttendanceStatus.EARLY_LEAVE:
            # 上班正常用正常顏色，下班早退用異常顏色
            if record.check_in:
                fill = self._get_fill(self.color_logic.normal_in_color)
                if fill:
                    in_cell.fill = fill
            fill = self._get_fill(self.color_logic.abnormal_out_color)
            if fill:
                out_cell.fill = fill
        
        elif record.status in (AttendanceStatus.ABNORMAL, AttendanceStatus.ABSENT):
            if record.check_in:
                fill = self._get_fill(self.color_logic.abnormal_in_color)
                if fill:
                    in_cell.fill = fill
            if record.check_out:
                fill = self._get_fill(self.color_logic.abnormal_out_color)
                if fill:
                    out_cell.fill = fill
    
    def _apply_missing_punch_color(self, cell):
        """Apply missing punch color based on color_logic settings."""
        color = self.color_logic.missing_punch_color
        fill = self._get_fill(color)
        if fill:
            cell.fill = fill
            # 如果是深色背景，使用白色文字
            if self._is_dark_color(color):
                cell.font = Font(color='FFFFFF')
    
    def _apply_absent_color(self, cell):
        """Apply absent color based on color_logic settings."""
        color = self.color_logic.absent_color
        fill = self._get_fill(color)
        if fill:
            cell.fill = fill
            # 如果是深色背景，使用白色文字
            if self._is_dark_color(color):
                cell.font = Font(color='FFFFFF')
    
    def apply_custom_colors(
        self,
        ws,
        row: int,
        day: int,
        in_color: Optional[str],
        out_color: Optional[str]
    ):
        """
        Apply custom colors to specific cells.
        
        Args:
            ws: Worksheet
            row: Starting row for the staff member (in-row)
            day: Day number (1-31)
            in_color: 'green', 'red', or None
            out_color: 'green', 'red', or None
        """
        col = day + 1
        
        if in_color and in_color in self.COLORS:
            ws.cell(row, col).fill = self.COLORS[in_color]
        
        if out_color and out_color in self.COLORS:
            ws.cell(row + 1, col).fill = self.COLORS[out_color]
    
    def _apply_person_border(self, ws, in_row: int, out_row: int, last_col: int):
        """
        為每個人的兩列套用粗外框線。
        
        Args:
            ws: Worksheet
            in_row: 上班時間列 (第一列)
            out_row: 下班時間列 (第二列)
            last_col: 最後一欄的欄號
        """
        for col in range(1, last_col + 1):
            in_cell = ws.cell(in_row, col)
            out_cell = ws.cell(out_row, col)
            
            # 取得現有邊框設定
            in_existing = in_cell.border
            out_existing = out_cell.border
            
            # 上列：上邊粗線，左右根據位置
            in_cell.border = Border(
                top=self.THICK_SIDE,
                bottom=in_existing.bottom if in_existing else self.THIN_SIDE,
                left=self.THICK_SIDE if col == 1 else self.THIN_SIDE,
                right=self.THICK_SIDE if col == last_col else self.THIN_SIDE
            )
            
            # 下列：下邊粗線，左右根據位置
            out_cell.border = Border(
                top=out_existing.top if out_existing else self.THIN_SIDE,
                bottom=self.THICK_SIDE,
                left=self.THICK_SIDE if col == 1 else self.THIN_SIDE,
                right=self.THICK_SIDE if col == last_col else self.THIN_SIDE
            )
