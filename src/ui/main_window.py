"""
Main Window Module

PyQt6 implementation of the 3-column attendance converter UI.
Matches the specified layout from the UI screenshot.
"""

import sys
from pathlib import Path
from datetime import datetime, date
from typing import Optional, List, Tuple

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QListWidget, QCheckBox, QSpinBox, QPushButton, QLineEdit,
    QTimeEdit, QGroupBox, QFrame, QFileDialog, QMessageBox,
    QApplication, QMenuBar, QMenu, QSplitter
)
from PyQt6.QtCore import Qt, QTime
from ui.styles import ThemeManager
from PyQt6.QtGui import QAction, QActionGroup, QFont

from config.config_manager import ConfigManager, AppConfig, TimeRule, ColorLogic, OutputSettings
from domain.entities import MonthlyStats
from infrastructure.filename_parser import FilenameParser


class MainWindow(QMainWindow):
    """
    Main application window with 3-column layout.
    
    Layout:
    - Left: Internal/External staff lists
    - Center: Time settings, color logic checkboxes, rate settings
    - Right: Statistics, PDF option
    - Bottom: Action buttons
    """
    
    # Default staff CSV filename
    DEFAULT_STAFF_CSV = "內外勤人員名單.csv"
    
    def __init__(self):
        super().__init__()
        self.config_manager = ConfigManager()
        self.config = self.config_manager.load()
        
        self._init_ui()
        self._load_config_to_ui()
        self._connect_signals()
        
        # Auto-load staff list on startup
        self._auto_load_staff_list()
    
    def _init_ui(self):
        """Initialize the user interface."""
        self.setWindowTitle("701Client 出勤報表轉換器")
        self.setMinimumSize(1000, 720)
        self.resize(1000, 720)
        
        # Menu bar
        self._create_menu_bar()
        
        # Central widget
        central = QWidget()
        self.setCentralWidget(central)
        
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)
        
        # Top area: 2 columns
        content_layout = QHBoxLayout()
        content_layout.setSpacing(15)
        
        # Left panel - Settings (Time, Checkboxes, Rate, Stats, Output)
        left_panel = self._create_settings_panel()
        content_layout.addWidget(left_panel, stretch=3)
        
        # Right panel - Staff Lists (Full height)
        staff_panel = self._create_staff_panel()
        content_layout.addWidget(staff_panel, stretch=1)
        
        main_layout.addLayout(content_layout, stretch=1)
        
        # Bottom area - Buttons
        bottom_panel = self._create_bottom_panel()
        main_layout.addWidget(bottom_panel)
        
        # Apply dark theme styling
        self._apply_styles()
    
    def _create_menu_bar(self):
        """Create the menu bar."""
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("檔案")
        
        import_staff_action = QAction("匯入人員名單", self)
        import_staff_action.triggered.connect(self._on_import_staff)
        file_menu.addAction(import_staff_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("離開", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Theme menu
        theme_menu = menubar.addMenu("主題")
        theme_group = QActionGroup(self)
        theme_group.setExclusive(True)
        
        # Get available themes and create actions
        current_theme = self.config.ui_prefs.theme_name
        
        for theme_name in ThemeManager.get_available_themes():
            action = QAction(theme_name, self, checkable=True)
            action.setData(theme_name)
            if theme_name == current_theme:
                action.setChecked(True)
            
            action.triggered.connect(lambda checked, name=theme_name: self._on_switch_theme(name))
            theme_menu.addAction(action)
            theme_group.addAction(action)

        # Settings action (在「幫助」左邊)
        settings_action = QAction("設定", self)
        settings_action.triggered.connect(self._on_open_settings)
        menubar.addAction(settings_action)

        # Help menu
        help_menu = menubar.addMenu("幫助")
        
        export_names_action = QAction("匯出xlsx人員名單", self)
        export_names_action.triggered.connect(self._on_export_names_from_xlsx)
        help_menu.addAction(export_names_action)
    
    def _create_staff_panel(self) -> QWidget:
        """Create the right panel with staff lists."""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)
        
        # Staff list statistics (Moved to bottom)
        staff_stats_layout = QHBoxLayout()
        staff_stats_layout.addWidget(QLabel("人員名單:"))
        
        self.lbl_staff_internal = QLabel("內勤: 0")
        self.lbl_staff_internal.setStyleSheet(
            "background-color: #2d5a27; color: white; padding: 3px 8px; "
            "border-radius: 3px; font-size: 11px;"
        )
        staff_stats_layout.addWidget(self.lbl_staff_internal)
        
        self.lbl_staff_external = QLabel("外勤: 0")
        self.lbl_staff_external.setStyleSheet(
            "background-color: #5a4427; color: white; padding: 3px 8px; "
            "border-radius: 3px; font-size: 11px;"
        )
        staff_stats_layout.addWidget(self.lbl_staff_external)
        staff_stats_layout.addStretch()
        
        # Internal staff list
        internal_group = QGroupBox("內勤名單")
        internal_layout = QVBoxLayout(internal_group)
        self.internal_list = QListWidget()
        self.internal_list.setMinimumHeight(100)
        self.internal_list.setFont(QFont("Microsoft JhengHei", 12))
        internal_layout.addWidget(self.internal_list)
        layout.addWidget(internal_group, stretch=1)
        
        # External staff list
        external_group = QGroupBox("外勤名單")
        external_layout = QVBoxLayout(external_group)
        self.external_list = QListWidget()
        self.external_list.setMinimumHeight(100)
        self.external_list.setFont(QFont("Microsoft JhengHei", 12))
        external_layout.addWidget(self.external_list)
        layout.addWidget(external_group, stretch=1)

        # Add stats at bottom
        layout.addLayout(staff_stats_layout)
        
        return panel
    
    def _create_settings_panel(self) -> QWidget:
        """Create the left panel with settings."""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)
        
        # Time settings group
        time_group = QGroupBox("上下班時間設定")
        time_layout = QGridLayout(time_group)
        time_layout.setSpacing(8)
        
        # Internal check-in
        time_layout.addWidget(QLabel("內勤上班時間"), 0, 0)
        self.internal_in_start = QTimeEdit()
        self.internal_in_start.setDisplayFormat("HH:mm")
        self.internal_in_end = QTimeEdit()
        self.internal_in_end.setDisplayFormat("HH:mm")
        time_layout.addWidget(self.internal_in_start, 0, 1)
        lbl_tilde_1 = QLabel("~")
        lbl_tilde_1.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_tilde_1.setStyleSheet("font-size: 18px; font-weight: bold;")
        time_layout.addWidget(lbl_tilde_1, 0, 2)
        time_layout.addWidget(self.internal_in_end, 0, 3)
        
        # Internal check-out
        time_layout.addWidget(QLabel("內勤下班時間"), 0, 4)
        self.internal_out_start = QTimeEdit()
        self.internal_out_start.setDisplayFormat("HH:mm")
        self.internal_out_end = QTimeEdit()
        self.internal_out_end.setDisplayFormat("HH:mm")
        time_layout.addWidget(self.internal_out_start, 0, 5)
        lbl_tilde_2 = QLabel("~")
        lbl_tilde_2.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_tilde_2.setStyleSheet("font-size: 18px; font-weight: bold;")
        time_layout.addWidget(lbl_tilde_2, 0, 6)
        time_layout.addWidget(self.internal_out_end, 0, 7)
        
        # External check-in
        time_layout.addWidget(QLabel("外勤上班時間"), 1, 0)
        self.external_in_start = QTimeEdit()
        self.external_in_start.setDisplayFormat("HH:mm")
        self.external_in_end = QTimeEdit()
        self.external_in_end.setDisplayFormat("HH:mm")
        time_layout.addWidget(self.external_in_start, 1, 1)
        lbl_tilde_3 = QLabel("~")
        lbl_tilde_3.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_tilde_3.setStyleSheet("font-size: 18px; font-weight: bold;")
        time_layout.addWidget(lbl_tilde_3, 1, 2)
        time_layout.addWidget(self.external_in_end, 1, 3)
        
        # External check-out
        time_layout.addWidget(QLabel("外勤下班時間"), 1, 4)
        self.external_out_start = QTimeEdit()
        self.external_out_start.setDisplayFormat("HH:mm")
        self.external_out_end = QTimeEdit()
        self.external_out_end.setDisplayFormat("HH:mm")
        time_layout.addWidget(self.external_out_start, 1, 5)
        lbl_tilde_4 = QLabel("~")
        lbl_tilde_4.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_tilde_4.setStyleSheet("font-size: 18px; font-weight: bold;")
        time_layout.addWidget(lbl_tilde_4, 1, 6)
        time_layout.addWidget(self.external_out_end, 1, 7)
        
        layout.addWidget(time_group, stretch=0)
        
        # Middle row: Attendance Rate + Stats
        middle_layout = QHBoxLayout()
        middle_layout.setSpacing(15)
        
        # 1. Attendance rate settings (Left side of middle row)
        rate_group = QGroupBox("出席率設定")
        rate_layout = QHBoxLayout(rate_group)
        
        # Threshold input
        threshold_layout = QVBoxLayout()
        threshold_layout.setSpacing(5)
        threshold_row = QHBoxLayout()
        threshold_row.setSpacing(8)
        threshold_row.addWidget(QLabel("出席率最低標準"))
        self.rate_threshold = QSpinBox()
        self.rate_threshold.setRange(0, 100)
        self.rate_threshold.setValue(80)
        self.rate_threshold.setSuffix(" %")
        self.rate_threshold.setMinimumWidth(80)
        threshold_row.addWidget(self.rate_threshold)
        threshold_row.addStretch()
        threshold_layout.addLayout(threshold_row)
        threshold_layout.addStretch()
        rate_layout.addLayout(threshold_layout)
        
        # Color legend
        legend_group = QGroupBox("顏色顯示配置")
        legend_layout = QGridLayout(legend_group)
        
        legend_layout.addWidget(QLabel("<80%"), 0, 0)
        red_label = QLabel("紅色")
        red_label.setStyleSheet("background-color: #FF6B6B; padding: 5px; border-radius: 3px;")
        legend_layout.addWidget(red_label, 0, 1)
        
        legend_layout.addWidget(QLabel(">=80%\n<90%"), 1, 0)
        yellow_label = QLabel("黃色")
        yellow_label.setStyleSheet("background-color: #FFD700; padding: 5px; border-radius: 3px;")
        legend_layout.addWidget(yellow_label, 1, 1)
        
        legend_layout.addWidget(QLabel(">=90%\n<=100%"), 2, 0)
        green_label = QLabel("綠色")
        green_label.setStyleSheet("background-color: #90EE90; padding: 5px; border-radius: 3px;")
        legend_layout.addWidget(green_label, 2, 1)
        
        rate_layout.addWidget(legend_group)
        middle_layout.addWidget(rate_group, stretch=1)
        
        # 2. Statistics (Right side of middle row)
        stats_group = self._create_stats_group()
        middle_layout.addWidget(stats_group, stretch=1)
        
        layout.addLayout(middle_layout, stretch=1)
        
        # Output settings group - simplified with single path
        output_group = QGroupBox("輸出設定")
        output_layout = QGridLayout(output_group)
        output_layout.setSpacing(8)
        
        # Output path row
        output_layout.addWidget(QLabel("輸出路徑"), 0, 0)
        self.output_path_edit = QLineEdit()
        self.output_path_edit.setPlaceholderText("Attendance_{year}_{month}.xlsx")
        self.output_path_edit.setReadOnly(True)  # Read only, use browse button
        output_layout.addWidget(self.output_path_edit, 0, 1)
        
        self.btn_browse_output = QPushButton("瀏覽")
        self.btn_browse_output.setMaximumWidth(60)
        self.btn_browse_output.clicked.connect(self._on_browse_output)
        output_layout.addWidget(self.btn_browse_output, 0, 2)
        
        # Generate PDF checkbox and hint
        self.chk_generate_pdf = QCheckBox("生成PDF")
        output_layout.addWidget(self.chk_generate_pdf, 1, 0)
        
        hint_label = QLabel("支援: {year}, {month}")
        hint_label.setStyleSheet("color: #888; font-size: 10px;")
        output_layout.addWidget(hint_label, 1, 1)
        
        layout.addWidget(output_group)
        
        return panel
    
    def _create_stats_group(self) -> QGroupBox:
        """Create the statistics group box."""
        group = QGroupBox("統計資訊")
        layout = QGridLayout(group)
        layout.setSpacing(8)
        
        # Required work days
        layout.addWidget(QLabel("本月應出席天數"), 0, 0)
        self.lbl_required_days = QLabel("0")
        self.lbl_required_days.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_required_days.setStyleSheet(
            "background-color: #3a4ad9; color: white; padding: 5px; "
            "border-radius: 4px; min-width: 50px; font-weight: bold;"
        )
        layout.addWidget(self.lbl_required_days, 0, 1)
        
        # Holidays
        layout.addWidget(QLabel("本月國定假日天數"), 1, 0)
        self.lbl_holidays = QLabel("0")
        self.lbl_holidays.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_holidays.setStyleSheet(
            "background-color: #3a4ad9; color: white; padding: 5px; "
            "border-radius: 4px; min-width: 50px; font-weight: bold;"
        )
        layout.addWidget(self.lbl_holidays, 1, 1)
        
        # Total source count
        layout.addWidget(QLabel("來源表單總人數"), 2, 0)
        self.lbl_total_count = QLabel("0")
        self.lbl_total_count.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_total_count.setStyleSheet(
            "background-color: #3a4ad9; color: white; padding: 5px; "
            "border-radius: 4px; min-width: 50px; font-weight: bold;"
        )
        layout.addWidget(self.lbl_total_count, 2, 1)
        
        # Internal count
        layout.addWidget(QLabel("來源表單內勤人員數"), 3, 0)
        self.lbl_internal_count = QLabel("0")
        self.lbl_internal_count.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_internal_count.setStyleSheet(
            "background-color: #3a4ad9; color: white; padding: 5px; "
            "border-radius: 4px; min-width: 50px; font-weight: bold;"
        )
        layout.addWidget(self.lbl_internal_count, 3, 1)
        
        # External count
        layout.addWidget(QLabel("來源表單外勤人員數"), 4, 0)
        self.lbl_external_count = QLabel("0")
        self.lbl_external_count.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_external_count.setStyleSheet(
            "background-color: #3a4ad9; color: white; padding: 5px; "
            "border-radius: 4px; min-width: 50px; font-weight: bold;"
        )
        layout.addWidget(self.lbl_external_count, 4, 1)
        
        return group
    
    def _create_bottom_panel(self) -> QWidget:
        """Create the bottom panel with action buttons."""
        panel = QWidget()
        layout = QHBoxLayout(panel)
        layout.setContentsMargins(0, 10, 0, 0)
        
        layout.addStretch()
        
        # Import leave list button
        self.btn_import_leave = QPushButton("匯入請假名單")
        self.btn_import_leave.setMinimumWidth(120)
        layout.addWidget(self.btn_import_leave)
        
        # Select source file button
        self.btn_select_source = QPushButton("選擇檔案")
        self.btn_select_source.setMinimumWidth(100)
        layout.addWidget(self.btn_select_source)
        
        # Convert button
        self.btn_convert = QPushButton("轉換")
        self.btn_convert.setMinimumWidth(80)
        self.btn_convert.setStyleSheet(
            "background-color: #2ea043; color: white; font-weight: bold;"
        )
        layout.addWidget(self.btn_convert)
        
        return panel
    
    def _apply_styles(self):
        """Apply visual styles to the window using ThemeManager."""
        theme_name = self.config.ui_prefs.theme_name
        theme = ThemeManager.get_theme(theme_name)
        self.setStyleSheet(theme.stylesheet)

    def _on_switch_theme(self, theme_name: str):
        """Handle theme switching."""
        self.config.ui_prefs.theme_name = theme_name
        self.config_manager.save()
        self._apply_styles()
    
    def _on_open_settings(self):
        """Handle settings menu action - open settings dialog."""
        from ui.settings_dialog import SettingsDialog
        dialog = SettingsDialog(self.config, self)
        if dialog.exec():
            # Settings were saved by dialog, persist to file
            self.config_manager.save()
    
    def _connect_signals(self):
        """Connect UI signals to handlers."""
        # Button signals
        self.btn_import_leave.clicked.connect(self._on_import_leave)
        self.btn_select_source.clicked.connect(self._on_select_source)
        self.btn_convert.clicked.connect(self._on_convert)
        
        # Auto-save on settings change
        self.rate_threshold.valueChanged.connect(self._on_settings_changed)
        
        # Time edits
        for edit in [
            self.internal_in_start, self.internal_in_end,
            self.internal_out_start, self.internal_out_end,
            self.external_in_start, self.external_in_end,
            self.external_out_start, self.external_out_end
        ]:
            edit.timeChanged.connect(self._on_settings_changed)
        
        # Output settings
        self.chk_generate_pdf.stateChanged.connect(self._on_settings_changed)
    
    def _load_config_to_ui(self):
        """Load configuration values into UI controls."""
        config = self.config
        
        # Time rules
        self._set_time_edit(self.internal_in_start, config.time_rules.internal.in_start)
        self._set_time_edit(self.internal_in_end, config.time_rules.internal.in_end)
        self._set_time_edit(self.internal_out_start, config.time_rules.internal.out_start)
        self._set_time_edit(self.internal_out_end, config.time_rules.internal.out_end)
        
        self._set_time_edit(self.external_in_start, config.time_rules.external.in_start)
        self._set_time_edit(self.external_in_end, config.time_rules.external.in_end)
        self._set_time_edit(self.external_out_start, config.time_rules.external.out_start)
        self._set_time_edit(self.external_out_end, config.time_rules.external.out_end)
        
        # Rate threshold
        self.rate_threshold.setValue(config.ui_prefs.rate_threshold)
        
        # Output settings - build full path from dir + pattern
        output_dir = config.output_settings.output_dir or str(Path.cwd())
        filename = config.output_settings.filename_pattern
        self.output_path_edit.setText(str(Path(output_dir) / filename))
        self.chk_generate_pdf.setChecked(config.output_settings.generate_pdf)
    
    def _set_time_edit(self, edit: QTimeEdit, time_str: str):
        """Set a QTimeEdit value from a time string."""
        try:
            parts = time_str.split(':')
            edit.setTime(QTime(int(parts[0]), int(parts[1])))
        except (ValueError, IndexError):
            edit.setTime(QTime(9, 0))
    
    def _get_time_str(self, edit: QTimeEdit) -> str:
        """Get time string from a QTimeEdit."""
        return edit.time().toString("HH:mm")
    
    def _save_ui_to_config(self):
        """Save UI values to configuration."""
        config = self.config
        
        # Time rules
        config.time_rules.internal.in_start = self._get_time_str(self.internal_in_start)
        config.time_rules.internal.in_end = self._get_time_str(self.internal_in_end)
        config.time_rules.internal.out_start = self._get_time_str(self.internal_out_start)
        config.time_rules.internal.out_end = self._get_time_str(self.internal_out_end)
        
        config.time_rules.external.in_start = self._get_time_str(self.external_in_start)
        config.time_rules.external.in_end = self._get_time_str(self.external_in_end)
        config.time_rules.external.out_start = self._get_time_str(self.external_out_start)
        config.time_rules.external.out_end = self._get_time_str(self.external_out_end)
        
        # Rate threshold
        config.ui_prefs.rate_threshold = self.rate_threshold.value()
        
        # Output settings - parse path back into dir + filename
        output_path = Path(self.output_path_edit.text()) if self.output_path_edit.text() else Path.cwd() / "Attendance_{year}_{month}.xlsx"
        config.output_settings.output_dir = str(output_path.parent)
        config.output_settings.filename_pattern = output_path.name
        config.output_settings.generate_pdf = self.chk_generate_pdf.isChecked()
        
        self.config_manager.save()
    
    def _on_settings_changed(self):
        """Handle settings change - save to config."""
        self._save_ui_to_config()
    
    def _show_message_box(self, msg_type: str, title: str, message: str):
        """Show a message box with black text color.
        
        Args:
            msg_type: Type of message box - 'information', 'warning', 'critical'
            title: Dialog title
            message: Message content
        """
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        
        # Set icon based on type
        if msg_type == "information":
            msg_box.setIcon(QMessageBox.Icon.Information)
        elif msg_type == "warning":
            msg_box.setIcon(QMessageBox.Icon.Warning)
        elif msg_type == "critical":
            msg_box.setIcon(QMessageBox.Icon.Critical)
        
        # Apply black text color style
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #ffffff;
            }
            QMessageBox QLabel {
                color: #000000;
                font-size: 13px;
            }
            QMessageBox QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
                min-width: 60px;
            }
            QMessageBox QPushButton:hover {
                background-color: #106ebe;
            }
        """)
        
        msg_box.exec()
    
    def _on_import_staff(self):
        """Handle import staff list action."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "選擇人員名單",
            "",
            "CSV Files (*.csv);;All Files (*)"
        )
        if file_path:
            self.config.paths.staff_csv = file_path
            self.config_manager.save()
            self._load_staff_list(Path(file_path))
    
    def _on_import_leave(self):
        """Handle import leave list button."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "選擇請假名單",
            "",
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        if file_path:
            self.config.paths.leave_list = file_path
            self.config_manager.save()
            self._show_message_box("information", "成功", f"已匯入請假名單：{file_path}")
    
    def _on_select_source(self):
        """Handle select source file button."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "選擇來源檔案",
            "",
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        if file_path:
            self.config.paths.last_source_file = file_path
            self.config_manager.save()
            self._process_source_file(Path(file_path))
    
    def _on_browse_output(self):
        """Handle browse output path button."""
        # Get current path as starting directory
        current_path = self.output_path_edit.text()
        if current_path:
            start_dir = str(Path(current_path).parent)
            default_filename = Path(current_path).name
        else:
            start_dir = str(Path.cwd())
            default_filename = "Attendance_{year}_{month}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "選擇輸出路徑",
            str(Path(start_dir) / default_filename),
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if file_path:
            self.output_path_edit.setText(file_path)
            self._save_ui_to_config()
    
    def _on_convert(self):
        """Handle convert button - generate report.
        
        Acts as the Controller: validates pre-conditions, coordinates the 
        report generation, and displays appropriate feedback to the user.
        """
        # Pre-condition Check 1: Source file
        source_path = self.config.paths.last_source_file
        if not source_path:
            self._show_message_box("warning", "警告", "請先選擇來源檔案")
            return
        
        # Pre-condition Check 2: Staff CSV (Strict Mode)
        staff_csv_path = self.config.paths.staff_csv
        if not staff_csv_path:
            self._show_message_box("warning", "警告", "未選取人員名單\n\n請先透過「檔案 → 匯入人員名單」設定人員名單。")
            return
        
        if not Path(staff_csv_path).exists():
            self._show_message_box(
                "warning",
                "警告", 
                f"未選取人員名單\n\n找不到人員名單檔案：\n{staff_csv_path}\n\n請重新匯入人員名單。"
            )
            return
        
        # Parse date from source filename to format output path
        from infrastructure.filename_parser import FilenameParser
        year_month = FilenameParser.try_parse_report_date(Path(source_path).name)
        
        if year_month:
            year, month = year_month
        else:
            # Fallback to current date if parsing fails
            now = datetime.now()
            year, month = now.year, now.month
            
        # Generate output path using config settings
        output_dir = self.config.output_settings.output_dir or str(Path.cwd())
        filename_pattern = self.config.output_settings.filename_pattern
        
        try:
            # Construct full path and format with year/month
            # Handle MM with leading zero
            full_pattern = Path(output_dir) / filename_pattern
            output_path = str(full_pattern).format(
                year=year, 
                month=f"{month:02d}"
            )
        except Exception as e:
            self._show_message_box("warning", "路徑格式錯誤", f"無法格式化輸出路徑：\n{str(e)}\n\n將使用預設檔名。")
            output_path = str(Path(output_dir) / f"Attendance_{year}_{month:02d}.xlsx")
            
        # Ensure output directory exists
        try:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            self._show_message_box("critical", "錯誤", f"無法建立輸出目錄：\n{str(e)}")
            return
        
        try:
            skipped_names = self._generate_report(Path(source_path), Path(output_path))
            
            # Feedback based on result
            if not skipped_names:
                # Scenario A: Perfect conversion
                self._show_message_box(
                    "information",
                    "成功", 
                    f"轉換成功！\n\n報表位於：\n{output_path}"
                )
            else:
                # Scenario B: Partial completion with warnings
                skipped_list = "\n".join(f"• {name}" for name in sorted(skipped_names))
                self._show_message_box(
                    "warning",
                    "轉換完成（含警告）",
                    f"報表已產生，但以下 {len(skipped_names)} 位人員不在人員名單中，已被略過：\n\n"
                    f"{skipped_list}\n\n"
                    f"報表位於：\n{output_path}"
                )
                
        except ValueError as e:
            # Scenario C: Complete failure (e.g., all staff skipped, no data)
            self._show_message_box("critical", "錯誤", f"轉換失敗：{str(e)}")
        except PermissionError:
            self._show_message_box(
                "critical",
                "錯誤", 
                f"無法寫入檔案，檔案可能被其他程式佔用：\n{output_path}"
            )
        except Exception as e:
            self._show_message_box("critical", "錯誤", f"生成報表時發生非預期錯誤：{str(e)}")
    
    def _on_export_names_from_xlsx(self):
        """Handle export names from xlsx - extract all names and save to txt.
        
        Creates a text file with all unique names from the source xlsx file.
        Each worksheet (sheet tab) name represents a person's name.
        Names are arranged left-to-right, then top-to-bottom in columns.
        """
        # Check if source file is selected
        source_path = self.config.paths.last_source_file
        if not source_path:
            self._show_message_box("warning", "警告", "請先選擇來源檔案")
            return
        
        if not Path(source_path).exists():
            self._show_message_box(
                "warning",
                "警告", 
                f"來源檔案不存在：\n{source_path}"
            )
            return
        
        try:
            from openpyxl import load_workbook
            
            # Load workbook and get all sheet names (each sheet = one person)
            wb = load_workbook(Path(source_path), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            # Filter out non-person sheets and clean names
            # Keywords that indicate system/summary sheets (not person names)
            excluded_keywords = (
                'sheet', 'summary', '匯總', '總表', '說明', 
                '出勤', '統計', '彙整', '工作表', '報表', '封面'
            )
            
            cleaned_names = []
            for sheet_name in sheet_names:
                # Skip sheets with excluded keywords
                name_lower = sheet_name.lower()
                if any(keyword in name_lower for keyword in excluded_keywords):
                    continue
                
                # Remove trailing '-' from name
                clean_name = sheet_name.rstrip('-')
                
                # Skip empty names after cleaning
                if clean_name:
                    cleaned_names.append(clean_name)
            
            if not cleaned_names:
                self._show_message_box("warning", "警告", "來源檔案中沒有找到任何人員名單")
                return
            
            # Sort names for better organization
            sorted_names = sorted(cleaned_names)
            
            # Calculate layout parameters
            # Assuming A4 page height ~50 lines at 12pt font, use 40 names per column as safe limit
            max_names_per_column = 40
            num_columns = (len(sorted_names) + max_names_per_column - 1) // max_names_per_column
            
            # Build the column layout
            columns = []
            for col_idx in range(num_columns):
                start_idx = col_idx * max_names_per_column
                end_idx = min(start_idx + max_names_per_column, len(sorted_names))
                columns.append(sorted_names[start_idx:end_idx])
            
            # Find the maximum column height
            max_height = max(len(col) for col in columns) if columns else 0
            
            # Build output lines (row by row across columns)
            output_lines = []
            column_width = 20  # Character width for each column
            
            for row_idx in range(max_height):
                row_parts = []
                for col in columns:
                    if row_idx < len(col):
                        name = col[row_idx]
                        # Left-align and pad to column width
                        row_parts.append(name.ljust(column_width))
                    else:
                        row_parts.append(' ' * column_width)
                # Join columns and strip trailing whitespace
                output_lines.append(''.join(row_parts).rstrip())
            
            # Generate output filename and path
            output_filename = "人員名單.txt"
            # Get project root directory (where main.py is located)
            project_root = Path(__file__).parent.parent.parent
            output_path = project_root / output_filename
            
            # Write to file with UTF-8 encoding
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(output_lines))
            
            self._show_message_box(
                "information",
                "成功", 
                f"人員名單已匯出！\n\n"
                f"共 {len(sorted_names)} 位人員\n"
                f"檔案位於：{output_path}"
            )
            
        except PermissionError:
            self._show_message_box(
                "critical",
                "錯誤", 
                "無法寫入檔案，檔案可能被其他程式佔用"
            )
        except Exception as e:
            self._show_message_box("critical", "錯誤", f"匯出人員名單時發生錯誤：{str(e)}")
    
    def _auto_load_staff_list(self):
        """Auto-load staff list on application startup.
        
        Priority:
        1. Check if config has a valid staff_csv path - use it if exists
        2. Try to load from default location (project root / DEFAULT_STAFF_CSV)
        3. If neither works, show warning dialog
        """
        # Check if config already has a valid staff CSV path
        config_csv_path = self.config.paths.staff_csv
        if config_csv_path and Path(config_csv_path).exists():
            try:
                self._load_staff_list(Path(config_csv_path))
                return  # Successfully loaded from config path
            except Exception:
                pass  # Config path failed, try default location
        
        # Try default location (same directory as main.py)
        project_root = Path(__file__).parent.parent.parent
        default_csv_path = project_root / self.DEFAULT_STAFF_CSV
        
        if default_csv_path.exists():
            try:
                self._load_staff_list(default_csv_path)
                # Update config with the default path
                self.config.paths.staff_csv = str(default_csv_path)
                self.config_manager.save()
                return  # Successfully loaded from default path
            except Exception as e:
                self._show_message_box(
                    "warning",
                    "警告",
                    f"載入預設人員名單時發生錯誤：\n{str(e)}\n\n請透過「檔案 → 匯入人員名單」手動添加。"
                )
                return
        
        # Neither path worked - show warning
        self._show_message_box(
            "warning",
            "找不到人員名單",
            f"找不到人員名單檔案：\n{self.DEFAULT_STAFF_CSV}\n\n"
            "請透過「檔案 → 匯入人員名單」手動添加人員名單。"
        )
    
    def _load_staff_list(self, csv_path: Path):
        """Load staff list from CSV and populate lists."""
        from domain.staff_classifier import StaffClassifier
        
        classifier = StaffClassifier()
        internal, external = classifier.load_from_csv(csv_path)
        
        self.internal_list.clear()
        for staff in internal:
            self.internal_list.addItem(staff.name)
        
        self.external_list.clear()
        for staff in external:
            self.external_list.addItem(staff.name)
        
        # Update staff list statistics (not source file stats)
        self.lbl_staff_internal.setText(f"內勤: {len(internal)}")
        self.lbl_staff_external.setText(f"外勤: {len(external)}")
    
    def _process_source_file(self, file_path: Path):
        """Process source file and update statistics."""
        from infrastructure.excel_parser import ExcelParser
        from infrastructure.filename_parser import FilenameParser
        from domain.rate_calculator import RateCalculator
        from domain.entities import Staff, StaffType
        from datetime import date
        
        parser = ExcelParser()
        
        # Try to extract year from filename (MonRepyymmdd format)
        parsed = FilenameParser.try_parse_report_date(file_path.name)
        year_from_filename = parsed[0] if parsed else None
        
        parser.parse_file(file_path, year=year_from_filename)
        
        names = parser.get_unique_names()
        
        # Update total count
        self.lbl_total_count.setText(str(len(names)))
        
        # Count internal/external based on loaded staff list
        internal_count = self.internal_list.count()
        external_count = self.external_list.count()
        
        self.lbl_internal_count.setText(str(internal_count))
        self.lbl_external_count.setText(str(external_count))
        
        # Parse year and month from filename (format: MonRepyymmdd)
        parsed = FilenameParser.try_parse_report_date(file_path.name)
        if parsed:
            year, month = parsed
            
            # Build holidays set from config
            holidays_set = set()
            for date_str in self.config.holidays.custom_dates:
                try:
                    # Expect format: YYYY-MM-DD
                    parts = date_str.split('-')
                    holidays_set.add(date(int(parts[0]), int(parts[1]), int(parts[2])))
                except (ValueError, IndexError):
                    pass
            
            # Create RateCalculator with holidays
            rate_calc = RateCalculator(holidays=holidays_set)
            
            # Create temporary internal staff to calculate required days
            # (UI displays internal staff work days as per requirement)
            temp_staff = Staff(name="temp", staff_type=StaffType.INTERNAL)
            required_days = rate_calc.calculate_required_days(temp_staff, year, month)
            
            # Calculate monthly stats to get holiday count
            stats = rate_calc.calculate_monthly_stats(year, month, internal_count, external_count)
            
            # Update UI
            self.lbl_required_days.setText(str(required_days))
            self.lbl_holidays.setText(str(stats.holidays))
    
    def _generate_report(self, source_path: Path, output_path: Path) -> List[str]:
        """Generate the attendance report with strict staff matching.
        
        This method implements strict matching logic:
        - Staff members NOT in the staff list will be SKIPPED (not processed).
        - Staff members in the list but NOT in source data are silently ignored.
        
        Args:
            source_path: Path to the source Excel file.
            output_path: Path to save the generated report.
            
        Returns:
            List[str]: A list of names that were skipped (found in source but not in staff list).
            
        Raises:
            ValueError: If no data found in source, or if all staff were skipped.
        """
        from infrastructure.excel_parser import ExcelParser
        from infrastructure.excel_writer import ExcelWriter
        from domain.staff_classifier import StaffClassifier
        from domain.attendance_logic import AttendanceLogicFactory
        from domain.rate_calculator import RateCalculator
        from domain.entities import StaffType, MonthlyAttendance, AttendanceRecord
        
        # Parse source file
        parser = ExcelParser()
        
        # Try to extract year from filename (MonRepyymmdd format)
        from infrastructure.filename_parser import FilenameParser
        parsed_date = FilenameParser.try_parse_report_date(source_path.name)
        year_from_filename = parsed_date[0] if parsed_date else None
        
        raw_data = parser.parse_file(source_path, year=year_from_filename)
        
        if not raw_data:
            raise ValueError("來源檔案中沒有找到任何資料")
        
        # Determine year and month from data
        first_date = raw_data[0].date
        year = first_date.year
        month = first_date.month
        
        # Get records by month
        records_by_name = parser.get_records_by_month(year, month)
        
        if not records_by_name:
            raise ValueError(f"來源檔案中沒有 {year} 年 {month} 月的資料")
        
        # Load staff classification (pre-condition already checked in _on_convert)
        classifier = StaffClassifier()
        classifier.load_from_csv(Path(self.config.paths.staff_csv))
        
        # Build holidays set from config (moved before loop to calculate work_days)
        holidays_set = set()
        for date_str in self.config.holidays.custom_dates:
            try:
                # Expect format: YYYY-MM-DD
                parts = date_str.split('-')
                holidays_set.add(date(int(parts[0]), int(parts[1]), int(parts[2])))
            except (ValueError, IndexError):
                pass
        
        # Get number of days in this month
        from calendar import monthrange
        _, num_days = monthrange(year, month)
        
        # Calculate attendance with strict matching
        rate_calc = RateCalculator()
        
        internal_attendance: List[MonthlyAttendance] = []
        external_attendance: List[MonthlyAttendance] = []
        skipped_names: List[str] = []
        
        for name, raw_rows in records_by_name.items():
            staff = classifier.get_staff_by_name(name)
            
            # Strict Matching: Skip if staff not in the list
            if not staff:
                skipped_names.append(name)
                continue
            
            # Convert to records
            records = parser.convert_to_attendance_records(raw_rows)
            
            # Apply logic
            strategy = AttendanceLogicFactory.get_strategy(staff.staff_type)
            time_rule = (
                self.config.time_rules.internal 
                if staff.staff_type == StaffType.INTERNAL 
                else self.config.time_rules.external
            )
            
            for record in records:
                record.status = strategy.determine_status(record, time_rule)
                record.remark = strategy.get_remark(record, time_rule)
            
            # Calculate work days for this staff member (based on staff type)
            work_days_set = set()
            for day in range(1, num_days + 1):
                d = date(year, month, day)
                if staff.should_work_on(d) and d not in holidays_set:
                    work_days_set.add(day)
            
            # Calculate monthly attendance with work_days filter
            monthly = rate_calc.calculate_monthly_attendance(
                staff, records, year, month,
                self.config.ui_prefs.rate_threshold,
                work_days=work_days_set
            )
            
            if staff.staff_type == StaffType.INTERNAL:
                internal_attendance.append(monthly)
            else:
                external_attendance.append(monthly)
        
        # Scenario C: Complete failure - all staff were skipped
        if not internal_attendance and not external_attendance:
            if skipped_names:
                raise ValueError(
                    f"所有 {len(skipped_names)} 位人員都不在人員名單中，無法產生報表。\n"
                    f"請確認人員名單是否正確。"
                )
            else:
                raise ValueError("沒有任何人員資料可供處理")
        
        # Apply sorting based on config setting
        from domain.sorting import sort_attendance_list
        sort_by = self.config.output_settings.sort_by
        internal_attendance = sort_attendance_list(internal_attendance, sort_by)
        external_attendance = sort_attendance_list(external_attendance, sort_by)
        
        # Generate Excel (holidays_set already built above)
        writer = ExcelWriter(self.config.ui_prefs.color_logic)
        writer.create_report(
            internal_attendance,
            external_attendance,
            year, month,
            output_path,
            holidays=holidays_set
        )
        
        # Generate PDF if checkbox is checked
        if self.chk_generate_pdf.isChecked():
            from infrastructure.pdf_writer import PdfWriter, format_filename
            
            pdf_writer = PdfWriter()
            
            # Determine PDF output directory
            pdf_output_dir = self.config.output_settings.pdf_output_dir
            if pdf_output_dir:
                pdf_dir = Path(pdf_output_dir)
            else:
                pdf_dir = output_path.parent
            
            # Ensure PDF output dir exists
            pdf_dir.mkdir(parents=True, exist_ok=True)
            
            # Check if separate or combined PDF
            if self.config.output_settings.separate_pdf:
                # Generate separate PDFs for internal and external
                if internal_attendance:
                    internal_pdf_pattern = self.config.output_settings.internal_pdf_pattern
                    internal_filename = format_filename(internal_pdf_pattern, year, month)
                    internal_pdf_path = pdf_dir / internal_filename
                    pdf_writer.create_report(
                        internal_attendance, year, month,
                        internal_pdf_path, "internal"
                    )
                
                if external_attendance:
                    external_pdf_pattern = self.config.output_settings.external_pdf_pattern
                    external_filename = format_filename(external_pdf_pattern, year, month)
                    external_pdf_path = pdf_dir / external_filename
                    pdf_writer.create_report(
                        external_attendance, year, month,
                        external_pdf_path, "external"
                    )
            else:
                # Generate combined PDF
                combined_pattern = self.config.output_settings.pdf_filename_pattern
                combined_filename = format_filename(combined_pattern, year, month)
                combined_pdf_path = pdf_dir / combined_filename
                pdf_writer.create_combined_report(
                    internal_attendance, external_attendance,
                    year, month, combined_pdf_path
                )
        
        return skipped_names
    
    def update_stats(self, stats: MonthlyStats):
        """Update statistics display."""
        self.lbl_required_days.setText(str(stats.required_work_days))
        self.lbl_holidays.setText(str(stats.holidays))
        self.lbl_total_count.setText(str(stats.total_staff_count))
        self.lbl_internal_count.setText(str(stats.internal_count))
        self.lbl_external_count.setText(str(stats.external_count))
    
    def closeEvent(self, event):
        """Handle window close - save config."""
        self._save_ui_to_config()
        event.accept()


def run_app():
    """Run the application."""
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    run_app()
