
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import unittest
from datetime import date, time
from unittest.mock import MagicMock
from openpyxl import Workbook

from infrastructure.excel_writer import ExcelWriter
from domain.entities import (
    Staff, StaffType, AttendanceRecord, AttendanceStatus, 
    MonthlyAttendance, RateColorTier
)
from config.config_manager import ColorLogic, TimeRule

class TestExcelWriterRemarks(unittest.TestCase):
    def setUp(self):
        self.writer = ExcelWriter()
        self.wb = Workbook()
        self.ws = self.wb.active

    def test_remarks_generation(self):
        # Setup mock data
        staff = Staff(name="Test User", staff_type=StaffType.INTERNAL)
        records = []
        
        # Day 1: Late (遲到)
        records.append(AttendanceRecord(
            date=date(2023, 10, 2), # Monday
            check_in=time(9, 10),
            check_out=time(18, 0),
            status=AttendanceStatus.LATE
        ))
        
        # Day 2: Early Leave (早退)
        records.append(AttendanceRecord(
            date=date(2023, 10, 3), # Tuesday
            check_in=time(9, 0),
            check_out=time(17, 30),
            status=AttendanceStatus.EARLY_LEAVE
        ))
        
        # Day 3: Overtime (超時打卡) - Status NORMAL but remark set
        records.append(AttendanceRecord(
            date=date(2023, 10, 4), # Wednesday
            check_in=time(9, 0),
            check_out=time(19, 0),
            status=AttendanceStatus.NORMAL,
            remark="下班延遲打卡"
        ))
        
        # Day 4: Normal
        records.append(AttendanceRecord(
            date=date(2023, 10, 5), # Thursday
            check_in=time(9, 0),
            check_out=time(18, 0),
            status=AttendanceStatus.NORMAL
        ))

        # Day 5: Late again
        records.append(AttendanceRecord(
            date=date(2023, 10, 6), # Friday
            check_in=time(9, 10),
            check_out=time(18, 0),
            status=AttendanceStatus.LATE
        ))
        
        monthly = MonthlyAttendance(
            staff=staff,
            year=2023,
            month=10,
            records=records,
            required_days=20,
            actual_days=20,
            attendance_rate=100.0,
            rate_color=RateColorTier.GREEN
        )

        # Execute
        self.writer._write_sheet(self.ws, [monthly], 2023, 10, is_external=False)

        # Verify
        # Expected work days in Oct 2023: Mon-Fri.
        # Columns start at 2 (1-based index). 
        # Oct 2023 starts on Sunday.
        # Work days: 2, 3, 4, 5, 6, 9, ...
        # Based on monthrange(2023, 10) -> (6, 31). 1st is Sun.
        # Work days = [2, 3, 4, 5, 6, 9...]. 
        # Total work days = 22 (approx). 
        # Remarks col = len(work_days) + 2.
        
        # Let's inspect the remarks cell directly by iterating cells in row 2
        # Remark cell should be the 3rd from last column written.
        
        # Easier way: iterate columns and find "備註" in row 1, then check row 2 same column
        remarks_col_idx = -1
        for cell in self.ws[1]:
            if cell.value == "備註":
                remarks_col_idx = cell.column
                break
        
        self.assertNotEqual(remarks_col_idx, -1, "Remarks column not found")
        
        remark_value = self.ws.cell(row=2, column=remarks_col_idx).value
        # Expected: 遲到2天, 早退1天, 超時打卡1天
        # Order depends on implementation: Late -> Early -> Overtime
        expected = "遲到2天, 早退1天, 超時打卡1天"
        self.assertEqual(remark_value, expected)

    def test_empty_remarks(self):
        staff = Staff(name="Perfect User", staff_type=StaffType.INTERNAL)
        records = [
            AttendanceRecord(
                date=date(2023, 10, 2),
                check_in=time(9, 0),
                check_out=time(18, 0),
                status=AttendanceStatus.NORMAL
            )
        ]
        monthly = MonthlyAttendance(staff, 2023, 10, records)
        
        self.writer._write_sheet(self.ws, [monthly], 2023, 10, is_external=False)
        
        remarks_col_idx = -1
        for cell in self.ws[1]:
            if cell.value == "備註":
                remarks_col_idx = cell.column
                break
                
        remark_value = self.ws.cell(row=2, column=remarks_col_idx).value
        self.assertEqual(remark_value, "")

if __name__ == '__main__':
    unittest.main()
